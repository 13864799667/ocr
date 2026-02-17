#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
批量文件夹处理脚本：自动处理多个年龄段文件夹的OCR识别

基于 one_key_process.py 的逻辑，扩展支持批量处理多个文件夹结构：
    大文件夹/
    ├── 20-女-166/
    │   └── 正常报告/
    │       └── 图/  ← 处理这里的图片
    ├── 30-女-166/
    │   └── 正常报告/
    │       └── 图/
    └── ...

主要功能：
1. 自动扫描所有年龄段文件夹
2. 依次处理每个 正常报告/图/ 目录
3. 每个年龄段独立处理，不累积数据
4. 自动打包 roi_images 和 all_color_output 为 zip
5. 将结果文件放到对应的 正常报告/ 目录

用法:
python batch_folder_process.py <大文件夹路径> [--roi_expand <ROI扩展像素>] [--debug]

示例:
python batch_folder_process.py /home/lys/ocr_demo/女-166 --debug

作者: Kiro
版本: v1.0
"""

import os
import sys
import argparse
import shutil
import zipfile
import time
import datetime
import logging
import re
import glob
import pandas as pd
import numpy as np
from PIL import Image, ImageEnhance, ImageFilter
from concurrent.futures import ProcessPoolExecutor, as_completed

# 修复numpy兼容性问题 - 在导入PaddleOCR之前
if not hasattr(np, 'int'):
    np.int = int
if not hasattr(np, 'float'):
    np.float = float
if not hasattr(np, 'bool'):
    np.bool = bool

from paddleocr import PaddleOCR
import cv2
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# ============================================================
# 以下函数直接复用 one_key_process.py 的核心逻辑
# ============================================================

def setup_logging(debug=False, log_file='batch_process.log'):
    """
    设置日志配置
    
    日志同时输出到:
    1. 控制台 - 实时查看进度
    2. 日志文件 - 保存完整记录，方便后续查错
    """
    level = logging.DEBUG if debug else logging.INFO
    
    # 清除现有的handlers
    root_logger = logging.getLogger()
    for handler in root_logger.handlers[:]:
        root_logger.removeHandler(handler)
    
    # 创建格式化器
    formatter = logging.Formatter(
        '%(asctime)s - %(levelname)s - %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    
    # 文件处理器 - 记录所有日志
    file_handler = logging.FileHandler(log_file, encoding='utf-8', mode='w')
    file_handler.setLevel(logging.DEBUG)  # 文件记录所有级别
    file_handler.setFormatter(formatter)
    
    # 控制台处理器 - 只显示INFO及以上
    console_handler = logging.StreamHandler()
    console_handler.setLevel(level)
    console_handler.setFormatter(formatter)
    
    # 配置根日志器
    root_logger.setLevel(logging.DEBUG)
    root_logger.addHandler(file_handler)
    root_logger.addHandler(console_handler)
    
    logging.info(f"日志文件: {log_file}")

def run_command(command):
    """运行命令行命令并返回结果"""
    import subprocess
    import platform
    
    logging.info(f"执行命令: {command}")
    try:
        if platform.system() == 'Windows':
            encoding = 'gbk'
        else:
            encoding = 'utf-8'
        
        result = subprocess.run(command, shell=True, check=True, 
                             stdout=subprocess.PIPE, stderr=subprocess.PIPE,
                             encoding=encoding, errors='ignore')
        return True, result.stdout
    except subprocess.CalledProcessError as e:
        error_msg = f"命令执行失败: {e.stderr if e.stderr else '未知错误'}"
        logging.error(error_msg)
        return False, error_msg
    except Exception as e:
        error_msg = f"命令执行异常: {str(e)}"
        logging.error(error_msg)
        return False, error_msg


def preprocess_image_for_ocr_v2(img, strategy='adaptive'):
    """
    增强版图像预处理函数，支持多种策略以提高OCR识别率
    （完全复用 one_key_process.py 的逻辑）
    """
    try:
        img_cv = cv2.cvtColor(np.array(img), cv2.COLOR_RGB2BGR)
        gray = cv2.cvtColor(img_cv, cv2.COLOR_BGR2GRAY)
        
        if strategy == 'adaptive':
            processed = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                                            cv2.THRESH_BINARY, 15, 5)
            kernel = np.ones((1, 1), np.uint8)
            processed = cv2.morphologyEx(processed, cv2.MORPH_CLOSE, kernel)
            
        elif strategy == 'contrast':
            clahe = cv2.createCLAHE(clipLimit=1.5, tileGridSize=(4,4))
            processed = clahe.apply(gray)
            _, processed = cv2.threshold(processed, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
            
        elif strategy == 'denoise':
            bilateral = cv2.bilateralFilter(gray, 5, 30, 30)
            blur = cv2.GaussianBlur(bilateral, (1, 1), 0)
            _, processed = cv2.threshold(blur, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
            
        elif strategy == 'sharpen':
            kernel = np.array([[-1,-1,-1], [-1,9,-1], [-1,-1,-1]])
            sharpened = cv2.filter2D(gray, -1, kernel)
            _, processed = cv2.threshold(sharpened, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
            
        elif strategy == 'negative':
            inverted = cv2.bitwise_not(gray)
            _, processed = cv2.threshold(inverted, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)

        elif strategy == 'enhanced':
            bilateral = cv2.bilateralFilter(gray, 5, 30, 30)
            clahe = cv2.createCLAHE(clipLimit=1.5, tileGridSize=(4,4))
            enhanced = clahe.apply(bilateral)
            processed = cv2.adaptiveThreshold(enhanced, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                                            cv2.THRESH_BINARY, 15, 5)
            kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, 1))
            processed = cv2.morphologyEx(processed, cv2.MORPH_CLOSE, kernel)

        elif strategy == 'minimal':
            processed = cv2.convertScaleAbs(gray, alpha=1.05, beta=3)

        elif strategy == 'gentle':
            enhanced = cv2.convertScaleAbs(gray, alpha=1.1, beta=5)
            processed = cv2.adaptiveThreshold(enhanced, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                                            cv2.THRESH_BINARY, 15, 5)

        elif strategy == 'original':
            processed = gray.copy()

        else:
            enhanced = cv2.convertScaleAbs(gray, alpha=1.1, beta=5)
            _, processed = cv2.threshold(enhanced, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
            kernel = np.ones((1, 1), np.uint8)
            processed = cv2.dilate(processed, kernel, iterations=1)
        
        processed_img = Image.fromarray(processed)
        
        if strategy in ['contrast', 'sharpen']:
            enhancer = ImageEnhance.Contrast(processed_img)
            processed_img = enhancer.enhance(1.5)
        
        return processed_img
        
    except Exception as e:
        logging.error(f"图像预处理失败 (策略: {strategy}): {str(e)}")
        return img

def extract_numbers_enhanced(text, allow_negative=True, feature_name=None):
    """增强版数字提取函数（完全复用 one_key_process.py 的逻辑）"""
    if not text:
        return ""

    cleaned_text = text.replace('O', '0').replace('o', '0').replace('I', '1').replace('l', '1')
    cleaned_text = cleaned_text.replace('，', '.').replace('。', '.').replace('·', '.')

    if feature_name:
        if "基础代谢" in feature_name:
            kcal_match = re.search(r'(\d+\.?\d*)\s*[kK][Cc][Aa][Ll]', cleaned_text)
            if kcal_match:
                value = float(kcal_match.group(1))
                return f"{value:.2f}"
            digits_match = re.search(r'(\d+\.?\d*)', cleaned_text)
            if digits_match:
                value = float(digits_match.group(1))
                return f"{value:.2f}"

        elif "健康分数" in feature_name:
            score_match = re.search(r'(\d+\.?\d*)', cleaned_text)
            if score_match:
                score = float(score_match.group(1))
                if score > 100:
                    score = score / 10
                return f"{score:.2f}"

    if allow_negative:
        negative_patterns = ['-', '—', '－', '―', '一']
        for pattern in negative_patterns:
            cleaned_text = cleaned_text.replace(pattern, '-')

    if allow_negative:
        number_pattern = r'[-]?\d*\.?\d+'
    else:
        number_pattern = r'\d*\.?\d+'

    numbers = re.findall(number_pattern, cleaned_text)

    valid_numbers = []
    for num in numbers:
        if num and num != '.' and num != '-':
            if re.search(r'\d', num):
                valid_numbers.append(num)

    if not valid_numbers:
        return ""

    return valid_numbers[0]

def expand_roi_coordinates(coords, img_width, img_height, expand_pixels=10):
    """智能扩展ROI坐标范围"""
    x1, y1, x2, y2 = coords
    new_x1 = max(0, x1 - expand_pixels)
    new_y1 = max(0, y1 - expand_pixels)
    new_x2 = min(img_width, x2 + expand_pixels)
    new_y2 = min(img_height, y2 + expand_pixels)
    return [new_x1, new_y1, new_x2, new_y2]

def process_muscle_data_enhanced(text, feature_name):
    """增强版节段肌肉数据处理函数"""
    if not text:
        return ""
    
    if "节段肌肉" not in feature_name:
        return text
    
    try:
        value = float(text)
        original_value = value
        
        if value > 1000:
            value = value / 1000
        elif value > 100:
            value = value / 100
        elif value > 50:
            value = value / 10
        
        if value > 50:
            value_str = str(int(original_value))
            if len(value_str) >= 2:
                value = float(value_str[:2])
            else:
                value = float(value_str)
        
        if value < 0.1:
            value = value * 10
        
        return f"{value:.2f}"
        
    except ValueError:
        digits = re.findall(r'\d+', text)
        if digits:
            try:
                value = float(digits[0])
                while value > 50:
                    value = value / 10
                while value < 0.1 and value > 0:
                    value = value * 10
                return f"{value:.2f}"
            except:
                pass
        return ""


def perform_ocr_paddle_enhanced(image_path, coords, feature_name, paddle_ocr, 
                              save_roi=True, roi_dir="roi_images", roi_expand=10):
    """增强版PaddleOCR识别函数（完全复用 one_key_process.py 的逻辑）"""
    try:
        img = Image.open(image_path)
        img_width, img_height = img.size
        
        expanded_coords = expand_roi_coordinates(coords, img_width, img_height, roi_expand)
        x1, y1, x2, y2 = expanded_coords
        
        roi = img.crop((x1, y1, x2, y2))
        
        if save_roi:
            if not os.path.exists(roi_dir):
                os.makedirs(roi_dir)
            
            base_name = os.path.basename(image_path)
            safe_feature_name = feature_name.replace('(','').replace(')','').replace('/','_').replace('-','_')
            roi_name = f"{os.path.splitext(base_name)[0]}_{safe_feature_name}.jpg"
            roi_path = os.path.join(roi_dir, roi_name)
            roi.save(roi_path)
        
        strategies = ['original', 'minimal', 'gentle', 'adaptive', 'contrast', 'enhanced']

        best_result = ""
        best_confidence = 0
        all_results = []

        for strategy in strategies:
            try:
                processed_roi = preprocess_image_for_ocr_v2(roi, strategy)

                if save_roi:
                    processed_roi_path = os.path.join(roi_dir, f"{strategy}_{roi_name}")
                    processed_roi.save(processed_roi_path)

                temp_path = os.path.join(roi_dir, f"temp_roi_{strategy}.jpg")
                processed_roi.save(temp_path)

                result = paddle_ocr.ocr(temp_path, cls=False)

                text = ""
                confidence = 0
                if result and len(result) > 0 and result[0] is not None:
                    for line in result[0]:
                        if len(line) >= 2:
                            text += line[1][0]
                            if isinstance(line[1], list) and len(line[1]) > 1:
                                confidence = max(confidence, line[1][1])

                if os.path.exists(temp_path):
                    os.remove(temp_path)

                if text.strip():
                    all_results.append((text, confidence, strategy))

                if text and confidence > best_confidence:
                    best_result = text
                    best_confidence = confidence

                if confidence > 0.9:
                    break

            except Exception as e:
                logging.debug(f"策略 '{strategy}' 识别失败: {str(e)}")
                continue

        if best_confidence < 0.5 and all_results:
            all_results.sort(key=lambda x: x[1], reverse=True)
            best_result = all_results[0][0]
            best_confidence = all_results[0][1]
        
        if best_result:
            allow_negative = any(keyword in feature_name.lower() for keyword in ['评估', '分数', '比'])
            cleaned_text = extract_numbers_enhanced(best_result, allow_negative, feature_name)
            
            if "节段肌肉" in feature_name:
                cleaned_text = process_muscle_data_enhanced(cleaned_text, feature_name)
            
            logging.info(f"OCR识别成功: {feature_name} = '{cleaned_text}'")
            return cleaned_text
        else:
            logging.warning(f"OCR识别失败: {feature_name}")
            return ""
            
    except Exception as e:
        logging.error(f"OCR识别错误: {feature_name} - {str(e)}")
        return ""

def load_coordinate_data_enhanced(coordinate_data):
    """增强版坐标数据加载函数"""
    coordinates = {}
    
    for line_num, line in enumerate(coordinate_data.strip().split('\n'), 1):
        line = line.strip()
        
        if not line or line.startswith('#'):
            continue
            
        parts = line.split()
        if len(parts) < 3:
            continue
            
        img_suffix = parts[0]
        feature_name = parts[1]
        
        coord_str = ' '.join(parts[2:])
        coords = re.findall(r'\d+', coord_str)
        
        if len(coords) < 4:
            continue
            
        try:
            x1, y1, x2, y2 = map(int, coords[:4])
            
            if x1 >= x2 or y1 >= y2:
                continue
                
            if img_suffix not in coordinates:
                coordinates[img_suffix] = []
                
            coordinates[img_suffix].append((feature_name, [x1, y1, x2, y2]))
            
        except ValueError:
            continue
    
    return coordinates

def find_original_file(base_name, original_dir):
    """根据分割后的目录名查找原始JPG文件"""
    if not original_dir or not os.path.exists(original_dir):
        return None, None
    
    try:
        possible_files = []
        
        for file in os.listdir(original_dir):
            file_path = os.path.join(original_dir, file)
            if os.path.isfile(file_path) and file.lower().endswith(('.jpg', '.jpeg', '.png')):
                if base_name in file or base_name in os.path.splitext(file)[0]:
                    possible_files.append(file_path)
        
        if possible_files:
            file_path = possible_files[0]
            mod_time = os.path.getmtime(file_path)
            mod_time_str = datetime.datetime.fromtimestamp(mod_time).strftime('%Y-%m-%d %H:%M:%S')
            return file_path, mod_time_str
            
    except Exception as e:
        pass
    
    return None, None

def process_images_paddle_enhanced(base_dir, coordinates, paddle_ocr, save_roi=True, 
                                 roi_dir="roi_images", original_dir=None, roi_expand=10):
    """单进程图片处理函数"""
    if not os.path.exists(base_dir):
        logging.error(f"目录 {base_dir} 不存在")
        return None
    
    results = {}
    
    try:
        image_files = [f for f in os.listdir(base_dir) if f.lower().endswith(('.jpg', '.jpeg', '.png'))]
    except Exception as e:
        logging.error(f"无法读取目录 {base_dir}，原因: {str(e)}")
        return None
    
    if not image_files:
        return None
    
    base_names = set()
    for img_file in image_files:
        parts = img_file.split('_part_')
        if len(parts) > 1:
            base_names.add(parts[0])
    
    if not base_names:
        return None
        
    for base_name in base_names:
        logging.info(f"处理文件: {base_name}")
        
        _, file_mod_time = find_original_file(base_name, original_dir)
        
        file_roi_dir = os.path.join(roi_dir, base_name) if save_roi else None
        if save_roi and not os.path.exists(file_roi_dir):
            os.makedirs(file_roi_dir)
        
        results['文件名'] = base_name
        
        if file_mod_time:
            results['原文件修改时间'] = file_mod_time
        else:
            try:
                if base_name.isdigit() and len(base_name) >= 10:
                    timestamp = int(base_name) / 1000 if len(base_name) > 10 else int(base_name)
                    approx_time = datetime.datetime.fromtimestamp(timestamp).strftime('%Y-%m-%d %H:%M:%S')
                    results['原文件修改时间'] = approx_time + " (估计)"
                else:
                    results['原文件修改时间'] = "未知"
            except:
                results['原文件修改时间'] = "未知"
        
        success_count = 0
        total_count = 0
        
        for img_suffix, features in coordinates.items():
            matching_files = [f for f in image_files if f.endswith(f"{img_suffix}.jpg")]
            if matching_files:
                img_path = os.path.join(base_dir, matching_files[0])
                for feature_name, coords in features:
                    total_count += 1
                    try:
                        text = perform_ocr_paddle_enhanced(img_path, coords, feature_name, 
                                                         paddle_ocr, save_roi, file_roi_dir, roi_expand)
                        results[feature_name] = text
                        if text:
                            success_count += 1
                    except Exception as e:
                        logging.error(f"处理特征 {feature_name} 时出错: {str(e)}")
                        results[feature_name] = ""
        
        if total_count > 0:
            success_rate = (success_count / total_count) * 100
            results['识别成功率'] = f"{success_rate:.1f}%"
    
    return results


def process_single_dir_worker(dir_path, coordinates, save_roi, roi_dir, original_dir, roi_expand):
    """
    单个目录处理的工作函数，用于并行处理
    （复用 one_key_process.py 的逻辑）
    
    参数:
    dir_path: 目录路径
    coordinates: 坐标字典
    save_roi: 是否保存ROI图像
    roi_dir: ROI图像保存目录
    original_dir: 原始JPG文件所在目录
    roi_expand: ROI扩展像素数

    返回:
    处理结果字典，如果失败返回None
    """
    try:
        # 检查是否有图片文件
        try:
            image_files = [f for f in os.listdir(dir_path) if f.lower().endswith(('.jpg', '.jpeg', '.png'))]
        except Exception as e:
            return None

        if not image_files:
            return None

        # 在工作进程中创建独立的PaddleOCR实例
        paddle_ocr = PaddleOCR(use_angle_cls=True, lang="ch", use_gpu=False,
                              show_log=False, det=True, rec_model_dir=None)

        results = process_images_paddle_enhanced(dir_path, coordinates, paddle_ocr, 
                                                 save_roi, roi_dir, original_dir, roi_expand)
        return results

    except Exception as e:
        logging.error(f"处理目录 {dir_path} 时出错: {str(e)}")
        return None


def process_batch_parallel(base_dir, coordinates, save_roi=True, roi_dir="roi_images", 
                          original_dir=None, roi_expand=10, max_workers=12):
    """
    并行批量处理函数
    （复用 one_key_process.py 的 process_batch_paddle_enhanced 逻辑）
    
    参数:
    base_dir: 包含分割后图片的目录
    coordinates: 坐标字典
    save_roi: 是否保存ROI图像
    roi_dir: ROI图像保存目录
    original_dir: 原始JPG文件所在目录
    roi_expand: ROI扩展像素数
    max_workers: 并行进程数
    
    返回:
    所有处理结果的列表
    """
    all_results = []

    # 获取所有子目录
    all_dirs = [base_dir]
    for root, dirs, files in os.walk(base_dir):
        for dir_name in dirs:
            all_dirs.append(os.path.join(root, dir_name))

    logging.info(f"找到 {len(all_dirs)} 个目录待处理")

    # 筛选出包含图片文件的目录
    valid_dirs = []
    for dir_path in all_dirs:
        try:
            image_files = [f for f in os.listdir(dir_path) if f.lower().endswith(('.jpg', '.jpeg', '.png'))]
            if image_files:
                valid_dirs.append(dir_path)
        except Exception as e:
            continue

    logging.info(f"找到 {len(valid_dirs)} 个包含图片的目录，使用 {max_workers} 个进程并行处理")
    processed_count = 0

    # 使用多进程并行处理目录
    with ProcessPoolExecutor(max_workers=max_workers) as executor:
        # 提交所有任务
        futures = [executor.submit(process_single_dir_worker, dir_path, coordinates,
                                  save_roi, roi_dir, original_dir, roi_expand)
                  for dir_path in valid_dirs]

        # 收集结果
        for future in as_completed(futures):
            try:
                results = future.result()
                if results:
                    all_results.append(results)
                    processed_count += 1
            except Exception as e:
                logging.error(f"并行处理任务失败: {str(e)}")

    logging.info(f"并行处理完成，成功处理 {processed_count} 个目录")
    return all_results


def save_excel_with_style_enhanced(df, output_file):
    """增强版Excel保存函数"""
    try:
        for col in df.columns:
            if col not in ['文件名', '原文件修改时间', '识别成功率']:
                df[col] = pd.to_numeric(df[col], errors='ignore')
        
        temp_excel = "temp_" + os.path.basename(output_file)
        df.to_excel(temp_excel, index=False)
        
        wb = openpyxl.load_workbook(temp_excel)
        ws = wb.active
        
        header_font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        data_font = Font(name='微软雅黑', size=11)
        data_alignment = Alignment(horizontal='center', vertical='center')
        
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = header_alignment
        
        for row in range(2, ws.max_row + 1):
            fill = PatternFill(start_color='EBF1F5', end_color='EBF1F5', fill_type='solid') if row % 2 == 0 else PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
            
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = data_alignment
                cell.fill = fill
        
        for col in range(1, ws.max_column + 1):
            max_length = 0
            column = get_column_letter(col)
            
            for row in range(1, ws.max_row + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
        
            adjusted_width = min(max_length + 4, 30)
            ws.column_dimensions[column].width = adjusted_width
        
        ws.freeze_panes = "A2"
        
        wb.save(output_file)
        
        if os.path.exists(temp_excel):
            os.remove(temp_excel)
            
        logging.info(f"Excel文件保存成功: {output_file}")
            
    except Exception as e:
        logging.warning(f"增强Excel格式化失败，使用基本格式保存: {str(e)}")
        df.to_excel(output_file, index=False)


# ============================================================
# 批量处理的核心逻辑
# ============================================================

# 坐标数据（与 one_key_process.py 保持一致）
COORDINATE_DATA_ENHANCED = """
_part_2 体重(kg) 41,57，139,107	
_part_3 体脂率(%) 38,27，136,75	
_part_4 骨骼肌(kg) 39,26，142,76	
_part_5 水分含量(kg) 169,96，236,125	
_part_7 浮肿评估 160,17，216,48	
_part_5 蛋白质(kg) 172,285，233,312
_part_5 无机盐(kg) 173,161，231,188
_part_11 BMI 156,25，219,64	
_part_13 节段肌肉(kg)-左臂 250,135，315,160	
_part_13 节段肌肉(kg)-右臂 44,133，127,160	
_part_13 节段肌肉(kg)-躯干 152,391，217,418	
_part_13 节段肌肉(kg)-左腿 249,291，315,320
_part_13 节段肌肉(kg)-右腿 46,288，127,319	
_part_15 腰臀比 157,25，219,63	
_part_16 内脏脂肪等级 161,27，216,62	
_part_17 基础代谢(kcal/day) 129,25，243,75
_part_18 健康分数 150,24，224,64
"""


def clear_directory(dir_path):
    """清空目录内容"""
    if os.path.exists(dir_path):
        try:
            shutil.rmtree(dir_path)
            logging.info(f"已清空目录: {dir_path}")
        except Exception as e:
            logging.warning(f"清空目录失败 {dir_path}: {str(e)}")
    os.makedirs(dir_path, exist_ok=True)


def zip_directory(source_dir, zip_path):
    """将目录打包为zip文件"""
    try:
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for root, dirs, files in os.walk(source_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    arcname = os.path.relpath(file_path, source_dir)
                    zipf.write(file_path, arcname)
        logging.info(f"打包完成: {zip_path}")
        return True
    except Exception as e:
        logging.error(f"打包失败 {zip_path}: {str(e)}")
        return False


def find_age_folders(base_path):
    """
    查找所有年龄段文件夹
    
    返回格式: [(年龄段名称, 正常报告/图 的完整路径), ...]
    """
    age_folders = []
    
    try:
        for item in sorted(os.listdir(base_path)):
            item_path = os.path.join(base_path, item)
            if os.path.isdir(item_path):
                # 检查是否存在 正常报告/图 目录
                normal_report_path = os.path.join(item_path, "正常报告", "图")
                if os.path.exists(normal_report_path) and os.path.isdir(normal_report_path):
                    # 检查是否有图片文件
                    image_files = [f for f in os.listdir(normal_report_path) 
                                   if f.lower().endswith(('.jpg', '.jpeg', '.png'))]
                    if image_files:
                        age_folders.append((item, normal_report_path))
                        logging.info(f"找到年龄段文件夹: {item} ({len(image_files)} 张图片)")
    except Exception as e:
        logging.error(f"扫描文件夹失败: {str(e)}")
    
    return age_folders


def process_single_age_folder(age_name, image_dir, coordinates, 
                              roi_expand=10, work_dir=".", max_workers=12):
    """
    处理单个年龄段文件夹
    
    参数:
    age_name: 年龄段名称（如 "20-女-166"）
    image_dir: 图片目录路径（正常报告/图）
    coordinates: 坐标字典
    roi_expand: ROI扩展像素数
    work_dir: 工作目录（脚本所在目录）
    max_workers: OCR并行进程数
    
    返回:
    (成功标志, 处理的文件数)
    """
    logging.info("=" * 60)
    logging.info(f"开始处理: {age_name}")
    logging.info("=" * 60)
    
    start_time = time.time()
    
    # 定义临时目录（在工作目录下）
    output_dir = os.path.join(work_dir, "all_color_output")
    roi_dir = os.path.join(work_dir, "roi_images")
    
    # 目标目录（正常报告文件夹）
    target_dir = os.path.dirname(image_dir)  # 正常报告/
    
    # 步骤1: 清空临时目录
    logging.info("步骤1: 清空临时目录")
    clear_directory(output_dir)
    clear_directory(roi_dir)
    
    # 步骤2: 执行图片分割（串行）
    logging.info("步骤2: 执行图片分割")
    split_script = os.path.join(work_dir, "split_image.py")
    if not os.path.exists(split_script):
        logging.error(f"分割脚本不存在: {split_script}")
        return False, 0
    
    input_pattern = os.path.join(image_dir, "*.jpg")
    split_cmd = f'python "{split_script}" --batch "{input_pattern}" "{output_dir}" auto color 5 0.995 5 80'
    
    success, output = run_command(split_cmd)
    if not success:
        logging.error(f"图片分割失败: {output}")
        return False, 0
    
    # 步骤3: 执行OCR识别（并行）
    logging.info(f"步骤3: 执行OCR识别（{max_workers}进程并行）")
    
    all_results = process_batch_parallel(
        output_dir, coordinates, 
        save_roi=True, roi_dir=roi_dir,
        original_dir=image_dir, roi_expand=roi_expand,
        max_workers=max_workers
    )
    
    if not all_results:
        logging.warning(f"没有识别到任何数据: {age_name}")
        return False, 0
    
    # 步骤4: 保存Excel
    logging.info("步骤4: 保存Excel")
    excel_filename = f"{age_name}_results.xlsx"
    excel_path = os.path.join(target_dir, excel_filename)
    
    df = pd.DataFrame(all_results)
    save_excel_with_style_enhanced(df, excel_path)
    
    # 步骤5: 打包roi_images
    logging.info("步骤5: 打包ROI图片")
    roi_zip_filename = f"{age_name}_roi_images.zip"
    roi_zip_path = os.path.join(target_dir, roi_zip_filename)
    zip_directory(roi_dir, roi_zip_path)
    
    # 步骤6: 打包all_color_output
    logging.info("步骤6: 打包分割输出")
    output_zip_filename = f"{age_name}_all_output.zip"
    output_zip_path = os.path.join(target_dir, output_zip_filename)
    zip_directory(output_dir, output_zip_path)
    
    # 步骤7: 清理临时目录
    logging.info("步骤7: 清理临时目录")
    clear_directory(output_dir)
    clear_directory(roi_dir)
    
    end_time = time.time()
    duration = end_time - start_time
    
    logging.info(f"✅ {age_name} 处理完成！")
    logging.info(f"   处理文件数: {len(all_results)}")
    logging.info(f"   耗时: {duration:.2f}秒")
    logging.info(f"   输出文件:")
    logging.info(f"     - {excel_path}")
    logging.info(f"     - {roi_zip_path}")
    logging.info(f"     - {output_zip_path}")
    
    return True, len(all_results)


def main():
    """主函数"""
    parser = argparse.ArgumentParser(
        description='批量文件夹处理脚本：自动处理多个年龄段文件夹的OCR识别',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  python batch_folder_process.py /home/lys/ocr_demo/女-166
  python batch_folder_process.py /home/lys/ocr_demo/女-166 --workers 8 --debug

文件夹结构要求:
  大文件夹/
  ├── 20-女-166/
  │   └── 正常报告/
  │       └── 图/  ← 处理这里的图片
  ├── 30-女-166/
  │   └── 正常报告/
  │       └── 图/
  └── ...

输出结果:
  每个年龄段的 正常报告/ 目录下会生成:
  - {年龄段}_results.xlsx      OCR识别结果
  - {年龄段}_roi_images.zip    ROI截图打包
  - {年龄段}_all_output.zip    分割图片打包
        """
    )
    parser.add_argument('base_dir', help='包含多个年龄段文件夹的大文件夹路径')
    parser.add_argument('--roi_expand', type=int, default=10, help='ROI扩展像素数 (默认: 10)')
    parser.add_argument('--workers', type=int, default=12, help='OCR并行进程数 (默认: 12)')
    parser.add_argument('--debug', action='store_true', help='启用调试模式')
    args = parser.parse_args()
    
    # 获取脚本所在目录作为工作目录
    work_dir = os.path.dirname(os.path.abspath(__file__))
    
    # 生成带时间戳的日志文件名
    timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    log_filename = f'batch_process_{timestamp}.log'
    log_file = os.path.join(work_dir, log_filename)
    
    # 设置日志
    setup_logging(args.debug, log_file)
    
    total_start_time = time.time()
    
    logging.info("=" * 70)
    logging.info("批量文件夹处理脚本 v1.1 (并行版)")
    logging.info(f"开始时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    logging.info(f"输入目录: {args.base_dir}")
    logging.info(f"工作目录: {work_dir}")
    logging.info(f"并行进程数: {args.workers}")
    logging.info("=" * 70)
    
    # 检查输入目录
    if not os.path.exists(args.base_dir):
        logging.error(f"输入目录不存在: {args.base_dir}")
        return
    
    # 检查必要的脚本文件
    split_script = os.path.join(work_dir, "split_image.py")
    if not os.path.exists(split_script):
        logging.error(f"分割脚本不存在: {split_script}")
        logging.error("请确保 split_image.py 在脚本同目录下")
        return
    
    # 查找所有年龄段文件夹
    logging.info("\n扫描年龄段文件夹...")
    age_folders = find_age_folders(args.base_dir)
    
    if not age_folders:
        logging.error("没有找到符合条件的年龄段文件夹")
        logging.error("请确保文件夹结构为: 年龄段/正常报告/图/")
        return
    
    logging.info(f"找到 {len(age_folders)} 个年龄段文件夹待处理")
    for age_name, image_dir in age_folders:
        logging.info(f"  - {age_name}")
    
    # 加载坐标数据
    logging.info("\n加载坐标数据...")
    coordinates = load_coordinate_data_enhanced(COORDINATE_DATA_ENHANCED)
    logging.info(f"已加载 {sum(len(features) for features in coordinates.values())} 个特征的坐标数据")
    
    # 处理统计
    success_count = 0
    fail_count = 0
    total_files = 0
    
    # 依次处理每个年龄段
    for idx, (age_name, image_dir) in enumerate(age_folders, 1):
        logging.info(f"\n{'='*70}")
        logging.info(f"进度: [{idx}/{len(age_folders)}]")
        
        try:
            success, file_count = process_single_age_folder(
                age_name, image_dir, coordinates,
                roi_expand=args.roi_expand, work_dir=work_dir,
                max_workers=args.workers
            )
            
            if success:
                success_count += 1
                total_files += file_count
            else:
                fail_count += 1
                
        except Exception as e:
            logging.error(f"处理 {age_name} 时发生错误: {str(e)}")
            fail_count += 1
            import traceback
            logging.debug(traceback.format_exc())
    
    # 最终清理
    logging.info("\n最终清理临时目录...")
    output_dir = os.path.join(work_dir, "all_color_output")
    roi_dir = os.path.join(work_dir, "roi_images")
    
    if os.path.exists(output_dir):
        shutil.rmtree(output_dir)
    if os.path.exists(roi_dir):
        shutil.rmtree(roi_dir)
    
    # 总结
    total_end_time = time.time()
    total_duration = total_end_time - total_start_time
    
    logging.info("\n" + "=" * 70)
    logging.info("批量处理完成！")
    logging.info("=" * 70)
    logging.info(f"处理结果:")
    logging.info(f"  - 成功: {success_count} 个年龄段")
    logging.info(f"  - 失败: {fail_count} 个年龄段")
    logging.info(f"  - 总文件数: {total_files} 个")
    logging.info(f"  - 总耗时: {total_duration:.2f}秒 ({total_duration/60:.1f}分钟)")
    if total_files > 0:
        logging.info(f"  - 平均每文件: {total_duration/total_files:.2f}秒")
    logging.info(f"结束时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    logging.info("=" * 70)
    
    # 控制台简洁输出
    print(f"\n🎉 批量处理完成！")
    print(f"   ✅ 成功: {success_count} 个年龄段")
    if fail_count > 0:
        print(f"   ❌ 失败: {fail_count} 个年龄段")
    print(f"   📊 总文件数: {total_files} 个")
    print(f"   ⏱️  总耗时: {total_duration:.2f}秒 ({total_duration/60:.1f}分钟)")
    print(f"   📝 详细日志: {log_file}")


if __name__ == "__main__":
    main()
